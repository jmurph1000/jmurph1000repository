from __future__ import annotations

import csv
import io
from typing import Any, Dict, List

from openpyxl import load_workbook


def read_spreadsheet_to_records(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    name = filename.lower()
    if name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]
    if name.endswith(".xlsx") or name.endswith(".xls"):
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        records: List[Dict[str, Any]] = []
        for r in rows[1:]:
            rec: Dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                rec[h] = r[i] if i < len(r) else None
            records.append(rec)
        return records
    raise ValueError("Unsupported file type. Please upload .csv or .xlsx")


def _normalize_keys(record: Dict[str, Any]) -> Dict[str, Any]:
    normalized: Dict[str, Any] = {}
    for k, v in record.items():
        if k is None:
            continue
        lc = str(k).strip().lower().replace(" ", "_")
        if lc in {"date", "as_of", "as_of_date"}:
            normalized["as_of_date"] = v
        elif lc in {"account", "account_name", "name"}:
            normalized["account_name"] = v
        elif lc in {"opening_balance", "balance", "amount", "opening"}:
            normalized["opening_balance"] = v
        elif lc in {"account_type", "type"}:
            normalized["account_type"] = v
        elif lc in {"type", "security_type", "asset_class"}:
            normalized["security_type"] = v
        elif lc in {"symbol", "ticker"}:
            normalized["symbol"] = v
        elif lc in {"qty", "quantity", "units"}:
            normalized["quantity"] = v
        elif lc in {"market_value", "mv", "value"}:
            normalized["market_value"] = v
        elif lc in {"ccy", "currency"}:
            normalized["currency"] = v
        else:
            normalized[lc] = v
    return normalized


def normalize_cash_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [_normalize_keys(r) for r in records]


def normalize_securities_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [_normalize_keys(r) for r in records]
